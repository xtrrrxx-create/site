"""
kiwifinder — Auto Import Script
================================
Importă produse din Excel/CSV direct în products.json.
Preia automat: imagine, preț, link kakobuy, link picksly.
Detectează automat categoria și batch-ul din titlu.

Rulare: deschide cu Python (IDLE → F5) sau dublu-click.
Fișierul auto_import.py trebuie să fie în același folder cu products.json.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json, os, re, threading, csv

# ── openpyxl opțional ──────────────────────────────────────────────────────
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── requests opțional (pentru fetch date live) ─────────────────────────────
try:
    import urllib.request, urllib.parse
    HAS_URLLIB = True
except ImportError:
    HAS_URLLIB = False

DB_FILE = "products.json"

# ══════════════════════════════════════════════════════════════════════════════
#  DETECTARE AUTOMATĂ
# ══════════════════════════════════════════════════════════════════════════════

CATEGORY_KEYWORDS = {
    "Shoes":       ["shoe", "shoes", "sneaker", "boot", "loafer", "jordan", "yeezy",
                    "nike", "adidas", "new balance", "puma", "samba", "dunk", "force",
                    "trail", "runner", "papuc", "pantofi", "incaltaminte"],
    "Shorts":      ["short", "shorts", "bermuda", "pantalon scurt"],
    "Pants":       ["pant", "pants", "trouser", "jean", "jeans", "denim", "cargo",
                    "sweatpant", "pantalon"],
    "T-shirts":    ["t-shirt", "tshirt", "tee", "tricou", "polo", "tank", "top"],
    "Hoodies":     ["hoodie", "hoody", "sweatshirt", "hanorac", "crewneck", "fleece",
                    "pullover", "zip-up"],
    "Jackets":     ["jacket", "coat", "parka", "windbreaker", "bomber", "geaca",
                    "varsity", "trench", "vest"],
    "Accessories": ["bag", "cap", "hat", "belt", "wallet", "watch", "sunglasses",
                    "glasses", "keychain", "sock", "socks", "scarf", "gloves",
                    "backpack", "crossbody", "geanta", "poseta", "sapca"],
}

BATCH_KEYWORDS = {
    "Best Batch":   ["vg", "og", "god tier", "best", "top", "premium", "lk", "pk",
                     "triple", "perfect", "rep 1:1", "1:1", "perfect rep"],
    "Budget Batch": ["budget", "cheap", "low", "basic", "economy"],
}

def detect_category(title: str) -> str:
    t = title.lower()
    for cat, kws in CATEGORY_KEYWORDS.items():
        for kw in kws:
            if kw in t:
                return cat
    return ""

def detect_batch(title: str) -> str:
    t = title.lower()
    for batch, kws in BATCH_KEYWORDS.items():
        for kw in kws:
            if kw in t:
                return batch
    return ""

# ══════════════════════════════════════════════════════════════════════════════
#  GENERARE LINKURI
# ══════════════════════════════════════════════════════════════════════════════

def make_kakobuy(url: str) -> str:
    if not url:
        return ""
    encoded = urllib.parse.quote(url, safe="")
    aff = os.environ.get("KAKOBUY_AFFCODE", "keviinn").strip() or "keviinn"
    return f"https://www.kakobuy.com/item/details?url={encoded}&affcode={aff}"

def _extract_first_url(text: str) -> str:
    if not text:
        return ""
    m = re.search(r'https?://[^\s",)]+', text, re.I)
    return m.group(0).strip() if m else ""

def _unwrap_kakobuy_url(url: str) -> str:
    """
    Primește URL kakobuy/ikako și încearcă să extragă link-ul sursă:
    - query params: url / itemUrl / target
    - redirect (pentru link-uri scurte ikako.vip)
    """
    if not url:
        return ""

    parsed = urllib.parse.urlparse(url)
    host = (parsed.netloc or "").lower()
    is_kako = any(h in host for h in ("kakobuy.com", "ikako.vip"))
    if not is_kako:
        return url

    qs = urllib.parse.parse_qs(parsed.query or "")
    for key in ("url", "itemUrl", "item_url", "target", "targetUrl", "redirect"):
        val = (qs.get(key) or [""])[0].strip()
        if val.startswith("http://") or val.startswith("https://"):
            return val

    if HAS_URLLIB:
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            with urllib.request.urlopen(req, timeout=12) as resp:
                final_url = resp.geturl() or ""
                if final_url and final_url != url:
                    return final_url
        except Exception:
            pass
    return url

def detect_source_platform(url: str) -> str:
    host = (urllib.parse.urlparse(url).netloc or "").lower()
    if "weidian.com" in host:
        return "weidian"
    if "taobao.com" in host or "tmall.com" in host:
        return "taobao"
    if "1688.com" in host or "alibaba.com" in host:
        return "1688"
    return ""

def make_picksly(url: str) -> str:
    """
    Picksly acceptă linkuri directe weidian/taobao.
    Formatul: https://picks.ly/item/<ITEM_ID_DIN_URL>
    Dacă nu poate extrage ID, returnează link-ul original.
    """
    if not url:
        return ""
    url = _unwrap_kakobuy_url(url)
    # Weidian: itemID=XXXXXXX
    m = re.search(r'itemID=(\d+)', url, re.I)
    if m:
        return f"https://picks.ly/item/WD{m.group(1)}"
    # Taobao: id=XXXXXXX
    m = re.search(r'[?&]id=(\d+)', url, re.I)
    if m:
        return f"https://picks.ly/item/TB{m.group(1)}"
    # 1688: /offer/XXXXXXX.html
    m = re.search(r'/offer/(\d+)\.html', url, re.I)
    if m:
        return f"https://picks.ly/item/1688{m.group(1)}"
    # Alibaba (unele linkuri sunt /product-detail/XXXX.html)
    m = re.search(r'/product-detail/[^/]*?(\d+)\.html', url, re.I)
    if m:
        return f"https://picks.ly/item/AL{m.group(1)}"
    return ""

def _extract_cny_price(url: str) -> str:
    """
    Extrage un preț CNY direct din HTML-ul paginii produsului.
    Returnează string gol dacă nu găsește nimic.
    """
    if not url or not HAS_URLLIB:
        return ""
    url = _unwrap_kakobuy_url(url)
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.google.com/"
        })
        with urllib.request.urlopen(req, timeout=10) as r:
            html = r.read().decode("utf-8", errors="ignore")

        patterns = [
            r'"price"\s*:\s*"(\d+(?:\.\d{1,2})?)"',
            r'"price"\s*:\s*(\d+(?:\.\d{1,2})?)',
            r'"priceText"\s*:\s*"[^"]*?(\d+(?:\.\d{1,2})?)"',
            r'"originPrice"\s*:\s*"(\d+(?:\.\d{1,2})?)"',
            r'¥\s?(\d+(?:\.\d{1,2})?)',
            r'￥\s?(\d+(?:\.\d{1,2})?)',
        ]
        for pat in patterns:
            m = re.search(pat, html, re.I)
            if m:
                return m.group(1)
    except Exception:
        pass
    return ""

def make_img(url: str) -> str:
    """
    Încearcă să preia URL-ul imaginii principale din pagina produsului.
    Dacă nu reușește, returnează string gol.
    """
    if not url or not HAS_URLLIB:
        return ""
    url = _unwrap_kakobuy_url(url)
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://weidian.com"
        })
        with urllib.request.urlopen(req, timeout=8) as r:
            html = r.read().decode("utf-8", errors="ignore")

        # Weidian — caută prima imagine din galerie
        patterns = [
            r'"pic":"(https?://[^"]+?\.jpg)"',
            r'"img":"(https?://[^"]+?\.jpg)"',
            r'<img[^>]+src="(https://si\.geilicdn\.com/[^"]+)"',
            r'"picList":\["(https?://[^"]+?\.jpg)"',
        ]
        for pat in patterns:
            m = re.search(pat, html)
            if m:
                img = m.group(1).replace("\\u002F", "/").replace("\\/", "/")
                return img
    except Exception:
        pass
    return ""

# ══════════════════════════════════════════════════════════════════════════════
#  CITIRE FIȘIER
# ══════════════════════════════════════════════════════════════════════════════

def read_file(path: str) -> list[dict]:
    """
    Acceptă Excel (.xlsx) sau CSV (.csv).
    Caută coloane: title/name, price/pret/cny, link/url/weidian/taobao,
                   img/image/imagine, kakobuy, picksly, category/categorie, batch
    Returnează listă de dict-uri cu cheile normalizate.
    """
    ext = os.path.splitext(path)[1].lower()

    rows = []
    headers = []

    if ext in (".xlsx", ".xls"):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl nu este instalat. Rulează: pip install openpyxl")
        # data_only=False ca să putem citi și formulele HYPERLINK(...)
        def cell_to_text(cell):
            # 1) hyperlink direct din celulă (Excel hyperlink object)
            try:
                if cell.hyperlink and getattr(cell.hyperlink, "target", None):
                    return str(cell.hyperlink.target).strip()
            except Exception:
                pass

            v = cell.value
            if v is None:
                return ""

            # 2) formulă HYPERLINK("url","text")
            if isinstance(v, str):
                m = re.search(r'HYPERLINK\(\s*"([^"]+)"', v, re.I)
                if m:
                    return m.group(1).strip()
                return v.strip()

            return str(v).strip()
        wb = openpyxl.load_workbook(path, data_only=False)

        # Alege automat worksheet-ul care pare listă de produse.
        best_ws = wb.active
        best_ws_header_idx = 0
        best_ws_score = -1
        for ws in wb.worksheets:
            all_rows_ws = list(ws.iter_rows())
            if not all_rows_ws:
                continue
            local_header_idx = 0
            local_score = -1
            for i, row_cells in enumerate(all_rows_ws[:40]):
                row_l = [cell_to_text(c).lower().strip() for c in row_cells]
                score = sum(1 for token in ("brand", "title", "name", "model", "link", "url", "price", "note")
                            if token in row_l)
                if score > local_score:
                    local_score = score
                    local_header_idx = i
            if local_score > best_ws_score:
                best_ws_score = local_score
                best_ws = ws
                best_ws_header_idx = local_header_idx

        all_rows = list(best_ws.iter_rows())
        if not all_rows:
            return []

        header_cells = all_rows[best_ws_header_idx]
        headers = [cell_to_text(c).lower().strip() for c in header_cells]

        for row_cells in all_rows[best_ws_header_idx + 1:]:
            values = [cell_to_text(c) for c in row_cells]
            if any(v for v in values):
                rows.append(dict(zip(headers, values)))

    elif ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            all_rows = list(csv.reader(f))
        if not all_rows:
            return []

        header_idx = 0
        score_best = -1
        for i, r in enumerate(all_rows[:40]):
            row_l = [str(c).lower().strip() for c in r]
            score = sum(1 for token in ("brand", "title", "name", "model", "link", "url", "price", "note")
                        if token in row_l)
            if score > score_best:
                score_best = score
                header_idx = i

        header = [str(c).strip() for c in all_rows[header_idx]]
        for raw in all_rows[header_idx + 1:]:
            if not any(str(c).strip() for c in raw):
                continue
            padded = list(raw) + [""] * max(0, len(header) - len(raw))
            row_dict = {}
            for k, v in zip(header, padded):
                key = str(k).lower().strip()
                row_dict[key] = str(v).strip()
            rows.append(row_dict)
    else:
        raise ValueError(f"Format nesuportat: {ext}. Folosește .xlsx sau .csv")

    return rows

def normalize_row(row: dict) -> dict:
    """Mapează orice nume de coloană la cheile standard."""
    def pick(*keys):
        for k in keys:
            for rk, rv in row.items():
                if k in rk and rv:
                    return rv
        return ""

    link_raw = pick("link", "url", "weidian", "taobao", "1688", "source", "kakobuy", "ikako")
    url_from_cell = _extract_first_url(link_raw)
    resolved_source = _unwrap_kakobuy_url(url_from_cell) if url_from_cell else ""
    platform = detect_source_platform(resolved_source)

    return {
        "title":    pick("title", "name", "nume", "produs", "denumire", "model", "item"),
        "price":    pick("price", "pret", "cny", "cost", "suma"),
        "link":     resolved_source,
        "img":      pick("img", "image", "imagine", "poza", "photo"),
        "kakobuy":  pick("kakobuy", "kako") or (make_kakobuy(resolved_source) if resolved_source else ""),
        "picksly":  pick("picksly", "picks", "qc") or (make_picksly(resolved_source) if resolved_source else ""),
        "category": pick("category", "categorie", "cat"),
        "batch":    pick("batch", "quality", "calitate", "tier"),
        "source_platform": platform,
    }

# ══════════════════════════════════════════════════════════════════════════════
#  DB
# ══════════════════════════════════════════════════════════════════════════════

def load_db() -> list:
    if not os.path.exists(DB_FILE):
        return []
    try:
        with open(DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def save_db(data: list):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

# ══════════════════════════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════════════════════════

class AutoImporter(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Kiwifinder — Auto Import")
        self.geometry("860x620")
        self.configure(bg="#121212")
        self.resizable(True, True)

        self.file_path = tk.StringVar()
        self.fetch_images = tk.BooleanVar(value=True)
        self.pending_rows: list[dict] = []  # rânduri normalizate din fișier

        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────
    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Treeview", background="#18181A", foreground="#fff",
                         fieldbackground="#18181A", rowheight=26)
        style.configure("Treeview.Heading", background="#27272A", foreground="white")
        style.map("Treeview", background=[("selected", "#3b82f6")])
        style.configure("TCheckbutton", background="#121212", foreground="white")

        # ── Titlu
        tk.Label(self, text="Auto Import Produse", font=("Inter", 18, "bold"),
                 bg="#121212", fg="white").pack(pady=(18, 4))
        tk.Label(self, text="Selectează un fișier Excel/CSV și importă automat în products.json",
                 font=("Arial", 9), bg="#121212", fg="#888").pack()

        # ── File picker
        fp_frame = tk.Frame(self, bg="#121212")
        fp_frame.pack(fill=tk.X, padx=24, pady=(14, 0))

        tk.Label(fp_frame, text="Fișier:", bg="#121212", fg="white",
                 font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Entry(fp_frame, textvariable=self.file_path, width=55,
                 bg="#18181A", fg="white", insertbackground="white",
                 relief="flat", bd=4).pack(side=tk.LEFT, padx=8)
        tk.Button(fp_frame, text="Browse", command=self._browse,
                  bg="#27272A", fg="white", relief="flat",
                  font=("Arial", 9, "bold"), padx=8).pack(side=tk.LEFT)

        # ── Opțiuni
        opt_frame = tk.Frame(self, bg="#121212")
        opt_frame.pack(fill=tk.X, padx=24, pady=6)
        ttk.Checkbutton(opt_frame, text="Preia imaginea automat din link (mai lent)",
                        variable=self.fetch_images,
                        style="TCheckbutton").pack(side=tk.LEFT)

        # ── Butoane acțiune
        btn_frame = tk.Frame(self, bg="#121212")
        btn_frame.pack(fill=tk.X, padx=24, pady=6)

        tk.Button(btn_frame, text="📂  Încarcă fișier", command=self._load_file,
                  bg="#27272A", fg="white", font=("Arial", 10, "bold"),
                  padx=12, pady=6, relief="flat").pack(side=tk.LEFT, padx=(0,8))
        tk.Button(btn_frame, text="✅  Importă în products.json", command=self._start_import,
                  bg="#3b82f6", fg="white", font=("Arial", 10, "bold"),
                  padx=12, pady=6, relief="flat").pack(side=tk.LEFT, padx=(0,8))
        tk.Button(btn_frame, text="🗑  Șterge selectate din preview", command=self._delete_selected,
                  bg="#ef4444", fg="white", font=("Arial", 10, "bold"),
                  padx=12, pady=6, relief="flat").pack(side=tk.LEFT)

        # ── Preview tabel
        cols = ("title", "price", "category", "batch", "kakobuy", "picksly", "img")
        col_labels = ("Titlu", "Preț", "Categorie", "Batch", "KakoBuy", "Picksly", "Img")
        col_widths  = (220, 60, 90, 100, 140, 140, 60)

        tree_frame = tk.Frame(self, bg="#121212")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=24, pady=(4, 0))

        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=14)
        for col, label, width in zip(cols, col_labels, col_widths):
            self.tree.heading(col, text=label)
            self.tree.column(col, width=width, anchor=tk.W)

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # ── Status bar
        self.status_var = tk.StringVar(value="Gata. Selectează un fișier Excel/CSV.")
        status_bar = tk.Frame(self, bg="#18181A", height=32)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)
        tk.Label(status_bar, textvariable=self.status_var,
                 bg="#18181A", fg="#aaa", font=("Arial", 9),
                 anchor="w", padx=10).pack(fill=tk.X)

        # ── Progress bar
        self.progress = ttk.Progressbar(self, mode="determinate", length=200)
        self.progress.pack(fill=tk.X, padx=24, pady=(4, 8))

    # ── Acțiuni ───────────────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Selectează fișier Excel sau CSV",
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("All", "*.*")]
        )
        if path:
            self.file_path.set(path)

    def _load_file(self):
        path = self.file_path.get().strip()
        if not path:
            messagebox.showwarning("Atenție", "Selectează mai întâi un fișier.")
            return
        try:
            raw_rows = read_file(path)
            self.pending_rows = [normalize_row(r) for r in raw_rows]

            # Șterge preview-ul vechi
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Populează preview cu detecție automată
            for r in self.pending_rows:
                cat   = r["category"] or detect_category(r["title"])
                batch = r["batch"]    or detect_batch(r["title"])
                kk    = r["kakobuy"]  or (make_kakobuy(r["link"]) if r["link"] else "")
                pk    = r["picksly"]  or (make_picksly(r["link"]) if r["link"] else "")
                img   = r["img"] or ("(se va prelua)" if self.fetch_images.get() and r["link"] else "")

                # Salvăm valorile completate înapoi în pending_rows
                r["_cat"]   = cat
                r["_batch"] = batch
                r["_kk"]    = kk
                r["_pk"]    = pk

                self.tree.insert("", tk.END, values=(
                    r["title"], r["price"], cat, batch,
                    kk[:40] + "..." if len(kk) > 40 else kk,
                    pk[:40] + "..." if len(pk) > 40 else pk,
                    "✓" if img else "✗"
                ))

            self.status_var.set(f"✓ {len(self.pending_rows)} produse încărcate din fișier. Verifică preview-ul și apasă Import.")
        except Exception as e:
            messagebox.showerror("Eroare", str(e))

    def _delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            return
        indices = sorted([self.tree.index(s) for s in selected], reverse=True)
        for idx in indices:
            self.tree.delete(self.tree.get_children()[idx])
            self.pending_rows.pop(idx)
        self.status_var.set(f"{len(self.pending_rows)} produse rămase în preview.")

    def _start_import(self):
        if not self.pending_rows:
            messagebox.showwarning("Atenție", "Încarcă mai întâi un fișier.")
            return
        # Rulează importul într-un thread ca să nu înghețe UI-ul
        threading.Thread(target=self._do_import, daemon=True).start()

    def _do_import(self):
        rows = self.pending_rows
        total = len(rows)
        self.progress["maximum"] = total
        self.progress["value"] = 0

        db = load_db()
        added = 0
        skipped = 0

        for i, r in enumerate(rows):
            title = r["title"]
            if not title:
                skipped += 1
                self._set_status(f"Omis rând fără titlu ({i+1}/{total})")
                self.progress["value"] = i + 1
                continue

            # Verifică duplicat după titlu
            if any(p.get("title", "").lower() == title.lower() for p in db):
                skipped += 1
                self._set_status(f"Duplicat omis: {title[:40]} ({i+1}/{total})")
                self.progress["value"] = i + 1
                continue

            cat   = r.get("_cat")   or detect_category(title)
            batch = r.get("_batch") or detect_batch(title)
            kk    = r.get("_kk")    or make_kakobuy(r["link"])
            pk    = r.get("_pk")    or make_picksly(r["link"])
            img   = r["img"]
            source_platform = detect_source_platform(r["link"])

            # Normalizează prețul CNY; dacă vine în alt format încearcă extragere din link.
            price = (r["price"] or "").strip()
            if price:
                cleaned = re.sub(r"[^0-9.]", "", price)
                price = cleaned if cleaned else price
            if (not price or "$" in r["price"]) and r["link"]:
                self._set_status(f"Extrage preț CNY: {title[:40]}... ({i+1}/{total})")
                live_price = _extract_cny_price(r["link"])
                if live_price:
                    price = live_price

            # Preia imaginea dacă e bifat și nu există deja
            if self.fetch_images.get() and not img and r["link"]:
                self._set_status(f"Preia imagine: {title[:40]}... ({i+1}/{total})")
                img = make_img(r["link"])

            product = {
                "title":    title,
                "price":    price,
                "img":      img,
                "kakobuy":  kk,
                "picksly":  pk,
                "category": cat,
                "batch":    batch,
                "source":   source_platform,
            }
            db.append(product)
            added += 1
            self._set_status(f"Adăugat: {title[:40]} ({i+1}/{total})")
            self.progress["value"] = i + 1

        try:
            save_db(db)
            self._set_status(f"✅ Import complet! {added} adăugate, {skipped} omise.")
            self.after(0, lambda: messagebox.showinfo(
                "Import complet",
                f"✅ {added} produse adăugate în products.json\n"
                f"⏭ {skipped} omise (duplicate sau fără titlu)"
            ))
        except PermissionError:
            self._set_status("❌ Eroare: nu am permisiune să scriu în products.json.")
            self.after(0, lambda: messagebox.showerror(
                "Eroare",
                "Nu am putut salva products.json.\n"
                "Asigură-te că auto_import.py este în același folder cu products.json\n"
                "și că rulezi din IDLE (nu dublu-click)."
            ))

    def _set_status(self, msg: str):
        self.after(0, lambda: self.status_var.set(msg))


if __name__ == "__main__":
    app = AutoImporter()
    app.mainloop()
